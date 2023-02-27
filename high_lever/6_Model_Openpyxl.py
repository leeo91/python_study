import openpyxl

# deal with Excel
# save data
workbook = openpyxl.Workbook()
sheet = workbook.create_sheet('weather')
sheet.append(['A', 'B', 'C'])
sheet.append(['zhangsan0', 'lisi', 'wangwu'])
workbook.save('weather.xlsx')


# read data
wk=openpyxl.load_workbook('weather.xlsx')
sheet=wk['weather']
lst=[]
for row in sheet.rows:
    rowlst=[]
    for cell in row:
        rowlst.append(cell.value)
    lst.append(rowlst)
for item in lst:
    print(item)
